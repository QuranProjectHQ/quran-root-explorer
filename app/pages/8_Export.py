# re-deploy 1779671310
"""Export page (WEB build) — three downloads streamed directly to browser.

Includes the Reading Guide (Interpret narrative) inside every format:
  • PDF   — first pages are the narrative, then every chart
  • HTML  — 00_Reading_Guide.html prepended to the zip
  • Excel — "Reading Guide" sheet inserted first

No server-side disk writes.  Everything streams to the visitor's browser.
"""
from __future__ import annotations

import io
import re
import zipfile

import streamlit as st
from PIL import Image

import analysis as A
import plotly_charts as PC
import stats_charts as SC
import stats_module as S
import interpret as I
from state import (get_corpus, query_controls, compute_all, need_results,
                   hero, layer, per_root_hint, log_page, log_export)

st.set_page_config(page_title="Export", page_icon="down", layout="wide")
log_page("export")
corpus = get_corpus()
raw, normalize, top_p, min_w, run = query_controls(corpus)
from state import needs_recompute
if run or needs_recompute():
    compute_all(corpus, raw, normalize, top_p, min_w)
R = need_results()

hero("Export",
     "PDF (universal) - HTML (interactive) - Excel (data). "
     "Reading Guide narrative is embedded in all three formats. "
     "Each one streams straight to your browser - no setup needed.")
per_root_hint(compact=True)


# ─────────────────────────────────────────────────────────────────
# Interpret narrative — generated once, embedded into every output
# ─────────────────────────────────────────────────────────────────
try:
    INTERP_SECTIONS = I.generate(R, corpus)
except Exception as _e:
    INTERP_SECTIONS = {"Reading Guide error": [f"{type(_e).__name__}: {str(_e)[:300]}"]}


def _strip_md(s: str) -> str:
    """Strip markdown bold markers and backticks for plain rendering."""
    return s.replace("**", "").replace("`", "")


def _interpret_html(sections: dict) -> str:
    parts = [
        "<!doctype html><html><head><meta charset='utf-8'>",
        "<title>Reading Guide — Quran Root Explorer</title>",
        "<style>",
        "body { font-family: -apple-system, Segoe UI, Roboto, sans-serif; "
        "max-width: 900px; margin: 2em auto; padding: 0 1em; color: #1B263B; "
        "line-height: 1.55; }",
        "h1 { color: #E63946; border-bottom: 3px solid #FCBF49; padding-bottom: 6px; }",
        "h2 { color: #1D3557; margin-top: 1.6em; "
        "background: linear-gradient(90deg,#FFF8E1,#FFFFFF); "
        "padding: 6px 10px; border-left: 5px solid #E63946; }",
        "ul { padding-left: 1.4em; }",
        "li { margin: 4px 0; }",
        "code { background: #FFF8E1; padding: 1px 6px; border-radius: 4px; "
        "color: #1B263B; font-family: 'Amiri', sans-serif; }",
        ".meta { color: #6B7280; font-size: 13px; }",
        "</style></head><body>",
        "<h1>Reading Guide</h1>",
        "<p class='meta'>Plain-English findings — every line is a number computed "
        "from your current session. No conjecture, no theological interpretation.</p>",
    ]
    for title, facts in sections.items():
        if not facts:
            continue
        parts.append(f"<h2>{title}</h2><ul>")
        for fact in facts:
            # convert **bold** → <b>, `code` → <code>
            html_fact = re.sub(r"\*\*(.*?)\*\*", r"<b>\1</b>", fact)
            html_fact = re.sub(r"`(.*?)`", r"<code>\1</code>", html_fact)
            parts.append(f"<li>{html_fact}</li>")
        parts.append("</ul>")
    parts.append("</body></html>")
    return "\n".join(parts)


def _interpret_pdf_figures(sections: dict):
    """Render the Reading Guide as a list of Plotly text-only figures, one per
    page. Each figure renders to PNG and is prepended to the PDF."""
    import plotly.graph_objects as _go
    LINES_PER_PAGE = 28
    pages: list[list[str]] = []
    current: list[str] = ["<b>READING GUIDE</b>",
                          "<i>Data-driven narrative from this session</i>", ""]
    for title, facts in sections.items():
        if not facts:
            continue
        block = [f"<b>{title}</b>"]
        for fact in facts:
            clean = _strip_md(fact)
            # Wrap long lines so the figure doesn't overflow
            if len(clean) > 100:
                clean = clean[:97] + "..."
            block.append(f"• {clean}")
        block.append("")
        if len(current) + len(block) > LINES_PER_PAGE and len(current) > 3:
            pages.append(current)
            current = []
        current.extend(block)
    if current:
        pages.append(current)

    figs = []
    for page_lines in pages:
        text = "<br>".join(page_lines)
        fig = _go.Figure()
        fig.update_layout(
            annotations=[dict(
                x=0.02, y=0.98, xref="paper", yref="paper",
                text=text, showarrow=False, align="left",
                font=dict(size=11, family="Arial"),
                xanchor="left", yanchor="top",
            )],
            xaxis=dict(visible=False),
            yaxis=dict(visible=False),
            margin=dict(l=20, r=20, t=20, b=20),
            paper_bgcolor="white", plot_bgcolor="white",
            showlegend=False, width=1280, height=720,
        )
        figs.append(fig)
    return figs


def _add_reading_guide_sheet(xlsx_bytes: bytes, sections: dict) -> bytes:
    """Open the built workbook, prepend a Reading Guide sheet, save back."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception:
        return xlsx_bytes
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    if "Reading Guide" in wb.sheetnames:
        del wb["Reading Guide"]
    ws = wb.create_sheet("Reading Guide", 0)
    ws["A1"] = "Reading Guide -- Quran Root Explorer"
    ws["A1"].font = Font(bold=True, size=14, color="E63946")
    ws["A2"] = ("Plain-English findings -- every line is a number computed from "
                "your current session. No conjecture.")
    ws["A2"].font = Font(italic=True, color="6B7280")
    row = 4
    section_fill = PatternFill("solid", fgColor="FFF8E1")
    for title, facts in sections.items():
        if not facts:
            continue
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12, color="1D3557")
        ws.cell(row=row, column=1).fill = section_fill
        row += 1
        for fact in facts:
            ws.cell(row=row, column=1, value=_strip_md(fact)).alignment = Alignment(wrap_text=True)
            row += 1
        row += 1
    ws.column_dimensions["A"].width = 110
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ─────────────────────────────────────────────────────────────────
# CHART CATALOG
# ─────────────────────────────────────────────────────────────────
def _build_catalog():
    g = R["graph"]
    comm = R["communities"]
    roots = R["input_roots"]
    catalog = []

    catalog += [
        ("01_Home", "01_summary.png",
         lambda: PC.chart_summary_metric_bars(R["summary"])),
        ("01_Home", "02_surah_distribution.png",
         lambda: PC.chart_distribution_across_surahs(R["occurrences"])),
        ("01_Home", "03_rarity_tier.png",
         lambda: PC.chart_rarity_tier(R["rarity"])),
    ]

    safe = lambda s: re.sub(r"[^\w\-]", "_", s)
    for r in roots:
        sub = f"02_PerRoot/{safe(r)}"
        catalog += [
            (sub, "01_surah_strip.png",
             lambda r=r: PC.chart_per_root_surah_strip(R["occurrences"], r)),
            (sub, "02_surface_forms.png",
             lambda r=r: PC.chart_surface_form_sunburst(R["sforms"], r)),
            (sub, "03_position_in_ayah.png",
             lambda r=r: PC.chart_position_histogram(R["position"], r)),
            (sub, "04_ayah_length.png",
             lambda r=r: PC.chart_ayah_length_hist(R["position"], r)),
            (sub, "05_partner_motifs.png",
             lambda r=r: PC.chart_partner_motifs(R["pmotifs"], r)),
        ]

    has_rev = R.get("has_rev_order", False)
    gm = R.get("g_meccan"); gd = R.get("g_medinan")
    dg = R.get("dg_lead_lag")
    catalog += [
        ("03_Network", "01_topology.png",
         lambda: PC.chart_network(g, comm)),
        ("03_Network", "02_chord_diagram.png",
         lambda: PC.chart_chord_diagram(g, comm)),
        ("03_Network", "03_adjacency_matrix.png",
         lambda: PC.chart_adjacency_matrix(g)),
        ("03_Network", "04_centrality_top20.png",
         lambda: PC.chart_centrality(R["centrality"], top=20)),
    ]
    if has_rev and gm is not None and gd is not None:
        def _build_stage(lo, hi):
            return A.build_phase_subgraph(corpus, roots, R["normalize"],
                                           lo, hi,
                                           top_partners=R["top_partners"],
                                           min_weight=R["min_weight"])
        catalog += [
            ("03_Network", "05_meccan_vs_medinan.png",
             lambda: PC.chart_phase_networks(gm, gd)),
            ("03_Network", "06_4stage_evolution.png",
             lambda: PC.chart_4stage_evolution(corpus, _build_stage)),
            ("03_Network", "07_phase_flow_sankey.png",
             lambda: PC.chart_sankey_phase_flow(gm, gd)),
            ("03_Network", "08_phase_diff.png",
             lambda: PC.chart_phase_diff_graph(gm, gd,
                                                R["phase_only_meccan"],
                                                R["phase_only_medinan"],
                                                R["phase_in_both"])),
        ]
    if dg is not None and dg.number_of_edges() > 0:
        catalog += [
            ("03_Network", "09_directed_lead_lag.png",
             lambda: PC.chart_directed_lead_lag(dg)),
            ("03_Network", "10_arc_diagram.png",
             lambda: PC.chart_arc_diagram(dg)),
        ]
    ns = R.get("net_stats", {})
    catalog += [
        ("03_Network", "11_ego_networks.png",
         lambda: PC.chart_per_root_ego_gallery(g, roots, max_neighbors=8)),
        ("03_Network", "12_robustness_overlay.png",
         lambda: PC.chart_robustness_overlay(g, ns.get("articulation_points", []),
                                              ns.get("bridge_edges", []), comm)),
        ("03_Network", "13_mst_backbone.png",
         lambda: PC.chart_mst_backbone(g)),
        ("03_Network", "14_kcore_layered.png",
         lambda: PC.chart_kcore_layered(g)),
        ("03_Network", "15_community_subnetworks.png",
         lambda: PC.chart_community_subnetworks(g, comm, top_n=12)),
        ("03_Network", "16_community_hierarchy.png",
         lambda: PC.chart_community_dendrogram(g, comm)),
    ]

    for size in (2, 3, 4):
        catalog.append(
            ("04_Motifs", f"0{size-1}_motif_gallery_size{size}.png",
             lambda size=size: PC.chart_motif_gallery(g, motif_size=size,
                                                       top_n=6,
                                                       input_roots=roots)))

    catalog += [
        ("05_Compare", "01_surah_heatmap.png",
         lambda: PC.chart_surah_heatmap(R["heatmap"])),
        ("05_Compare", "02_overlap_heatmap.png",
         lambda: PC.chart_overlap_heatmap(R["overlap"])),
        ("05_Compare", "03_pair_overlap_grouped.png",
         lambda: PC.chart_pair_overlap_grouped(R["overlap"],
                                                R["overlap_surah"],
                                                R["input_roots"])),
    ]

    catalog.append(("06_Morphology", "01_morphology_global.png",
                    lambda: PC.chart_morphology(R["morphology"])))
    for r in roots:
        catalog.append(
            (f"06_Morphology/{safe(r)}", "01_particles.png",
             lambda r=r: PC.chart_morphology_per_root(R["morphology"], r)))

    def _safe_stats(fn, *args, **kw):
        try: return fn(*args, **kw)
        except Exception: return None
    freq = _safe_stats(S.frequency_analysis, corpus, roots, R["normalize"])
    pos_df = _safe_stats(S.position_categorization, corpus, roots, R["normalize"])
    pmi_df = _safe_stats(S.pmi_matrix, corpus, roots, R["normalize"])
    cond_df = _safe_stats(S.conditional_probability, corpus, roots, R["normalize"])
    jac_df = _safe_stats(S.jaccard_matrix, corpus, roots, R["normalize"])
    surah_role = _safe_stats(S.surah_role, corpus, roots, R["normalize"])
    tfidf_df = _safe_stats(S.surah_tfidf, corpus, roots, R["normalize"])
    enr_df = _safe_stats(S.surah_enrichment, corpus, roots, R["normalize"])
    cum_df = _safe_stats(S.cumulative_trajectory, corpus, roots, R["normalize"])
    net_x = _safe_stats(S.network_extras, R["graph"])
    excl = _safe_stats(S.exclusive_partners, corpus, roots, R["normalize"])

    catalog += [
        ("07_Statistics", "01_frequency_bars.png",
         lambda: SC.chart_frequency_bars(freq)) if freq is not None else None,
        ("07_Statistics", "02_dispersion.png",
         lambda: SC.chart_dispersion(freq)) if freq is not None else None,
        ("07_Statistics", "03_position_tiles.png",
         lambda: SC.chart_position_tiles(pos_df)) if pos_df is not None else None,
        ("07_Statistics", "04_pmi.png",
         lambda: SC.chart_pmi_heatmap(pmi_df)) if pmi_df is not None else None,
        ("07_Statistics", "05_cond_prob.png",
         lambda: SC.chart_cond_prob_heatmap(cond_df)) if cond_df is not None else None,
        ("07_Statistics", "06_cond_prob_reverse.png",
         lambda: SC.chart_cond_prob_reverse_heatmap(cond_df)) if cond_df is not None else None,
        ("07_Statistics", "07_jaccard.png",
         lambda: SC.chart_jaccard_heatmap(jac_df)) if jac_df is not None else None,
        ("07_Statistics", "08_surah_role.png",
         lambda: SC.chart_surah_role_bar(surah_role)) if surah_role is not None else None,
        ("07_Statistics", "09_tfidf.png",
         lambda: SC.chart_tfidf_dot(tfidf_df)) if tfidf_df is not None else None,
        ("07_Statistics", "10_enrichment.png",
         lambda: SC.chart_enrichment_scatter(enr_df)) if enr_df is not None else None,
        ("07_Statistics", "11_cumulative.png",
         lambda: SC.chart_cumulative(cum_df)) if cum_df is not None else None,
        ("07_Statistics", "12_network_extras.png",
         lambda: SC.chart_network_extras(net_x)) if net_x is not None else None,
        ("07_Statistics", "13_exclusive_partners.png",
         lambda: SC.chart_exclusive_partners(excl)) if excl is not None else None,
        ("07_Statistics", "14_dendrogram.png",
         lambda: SC.chart_dendrogram(jac_df)) if jac_df is not None else None,
        ("07_Statistics", "15_metric_cross_reference.png",
         lambda: SC.chart_metric_cross_reference(pmi_df, jac_df, roots))
         if (pmi_df is not None and jac_df is not None) else None,
    ]

    # PREPEND Reading Guide pages — one Plotly text-figure per ~28-line page.
    interp_figs = _interpret_pdf_figures(INTERP_SECTIONS)
    interp_entries = [
        ("00_ReadingGuide", f"{i+1:02d}_page.png", (lambda f=f: f))
        for i, f in enumerate(interp_figs)
    ]
    return interp_entries + [x for x in catalog if x is not None]


catalog = _build_catalog()


# ─────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────
from collections import defaultdict
buckets = defaultdict(int)
for folder, _, _ in catalog:
    buckets[folder.split("/", 1)[0]] += 1

layer(1, "What you'll get")
c1, c2, c3 = st.columns(3)
c1.metric("Charts + narrative pages in PDF", len(catalog))
c2.metric("Files in HTML zip", len(catalog) + 1)  # +1 for 00_Reading_Guide.html
c3.metric("Sheets in Excel (incl. Reading Guide)", 14)
st.markdown(
    "All three artifacts include the **Reading Guide narrative** plus every chart "
    f"and table built from your input roots (`{' · '.join(R['input_roots']) or '—'}`). "
    "Everything streams straight to your browser's Downloads."
)

st.divider()

# Excel
meta = {
    "Input roots": " ".join(R["input_roots"]),
    "Normalization": "ON" if R["normalize"] else "OFF (exact)",
    "Co-occurrence scope": "same ayah",
    "Top partners": R["top_partners"],
    "Min edge weight": R["min_weight"],
    "Ayahs in corpus": corpus.n_ayahs,
    "Ayahs matched": len(R["match_ayahs"]),
}
_xlsx_buf = io.BytesIO()
A.export_excel(
    _xlsx_buf, summary=R["summary"], occurrences=R["occurrences"],
    cooccurrence_tbl=R["cooc_tbl"], surface_forms=R["sforms"],
    partner_motifs_tbl=R["pmotifs"], triangles_tbl=R["triangles"],
    triad_summary=R["triad"], meta=meta,
    centrality=R["centrality"], heatmap=R["heatmap"], overlap=R["overlap"],
    morphology=R["morphology"], position=R["position"], rarity=R["rarity"],
    first_last=R["first_last"],
)
_xlsx_bytes = _add_reading_guide_sheet(_xlsx_buf.getvalue(), INTERP_SECTIONS)


# Detect PDF renderer
def _probe_png():
    try:
        import plotly.graph_objects as _go
        _f = _go.Figure(_go.Scatter(x=[1, 2], y=[1, 2]))
        _f.to_image(format="png", width=200, height=150)
        return True, ""
    except Exception as _e:
        return False, f"{type(_e).__name__}: {str(_e)[:300]}"

png_ok, png_err = _probe_png()


layer(2, "Build PDF + interactive HTML")
st.caption(
    f"~1-3 seconds per chart x **{len(catalog)} items** ~= "
    f"**{round(len(catalog) * 2 / 60, 1)} minutes**. "
    "Click once and wait - the download buttons appear when ready."
)

if png_ok:
    st.success("PDF rendering is ready.")
else:
    st.warning(
        "PDF rendering is unavailable on this server right now - "
        "you can still get the interactive HTML zip and the Excel workbook."
    )
    with st.expander("Technical detail", expanded=False):
        st.code(png_err)

if st.button("Build PDF + HTML", type="primary",
             width='stretch', key="gen_all_charts"):
    progress = st.progress(0, text="Starting...")
    png_blobs = []
    html_blobs = []
    errors = []
    import plotly.io as _pio
    for i, (folder, fname, factory) in enumerate(catalog, 1):
        progress.progress(i / len(catalog),
                          text=f"{i}/{len(catalog)} - {folder}/{fname}")
        try:
            fig = factory()
            if fig is None:
                errors.append(f"{folder}/{fname}: factory returned None")
                continue
        except Exception as e:
            errors.append(f"{folder}/{fname}: factory() {type(e).__name__}: "
                          f"{str(e)[:300]}")
            continue
        try:
            html = _pio.to_html(fig, full_html=True, include_plotlyjs="cdn",
                                config={"displaylogo": False})
            html_blobs.append((folder, fname.replace(".png", ".html"), html))
        except Exception as e:
            errors.append(f"{folder}/{fname}: HTML {type(e).__name__}: "
                          f"{str(e)[:300]}")
        if png_ok:
            try:
                png = fig.to_image(format="png", width=1280, height=720, scale=1)
                png_blobs.append((folder, fname, png))
            except Exception as e:
                errors.append(f"{folder}/{fname}: PNG {type(e).__name__}: "
                              f"{str(e)[:300]}")
    progress.empty()

    if png_blobs:
        try:
            pdf_buf = io.BytesIO()
            imgs = [Image.open(io.BytesIO(b)).convert("RGB")
                    for (_, _, b) in png_blobs]
            if imgs:
                imgs[0].save(pdf_buf, format="PDF",
                             save_all=True, append_images=imgs[1:],
                             resolution=100.0)
                st.session_state["_export_pdf"] = pdf_buf.getvalue()
                st.session_state["_export_pdf_count"] = len(png_blobs)
        except Exception as e:
            st.warning(f"PDF assembly failed: {e}")

    if html_blobs:
        z = io.BytesIO()
        with zipfile.ZipFile(z, "w", zipfile.ZIP_DEFLATED) as zf:
            # Prepend the Reading Guide HTML as the first file in the zip
            zf.writestr("00_Reading_Guide.html", _interpret_html(INTERP_SECTIONS))
            for folder, fname, html in html_blobs:
                zf.writestr(f"{folder}/{fname}", html)
            seen, lines = set(), [
                "Quran Root Analysis - interactive charts + Reading Guide", "",
                f"Input roots: {' '.join(R['input_roots'])}",
                f"Total charts: {len(html_blobs)}", "",
                "Open 00_Reading_Guide.html first for the data-driven narrative.", "",
                "Folder structure:"]
            for folder, fname, _ in html_blobs:
                if folder not in seen:
                    lines.append(f"  {folder}/"); seen.add(folder)
                lines.append(f"    {fname}")
            zf.writestr("MANIFEST.txt", "\n".join(lines))
        st.session_state["_export_html_zip"] = z.getvalue()
        st.session_state["_export_html_count"] = len(html_blobs) + 1

    n_pdf = st.session_state.get("_export_pdf_count", 0)
    n_html = st.session_state.get("_export_html_count", 0)
    if n_pdf or n_html:
        st.success(
            f"Built **{n_pdf}** pages into PDF and **{n_html}** files in HTML zip "
            "(Reading Guide narrative is the first file in each). "
            "Download buttons are below."
        )

    if errors:
        with st.expander(f"{len(errors)} chart(s) had issues", expanded=False):
            for e in errors[:80]:
                st.code(e, language=None)


st.divider()
layer(3, "Download")

import hashlib
_roots_tag = "_".join(R["input_roots"][:3]) or "all"
_hash = hashlib.md5(("|".join(R["input_roots"])).encode("utf-8")).hexdigest()[:6]
_stem = f"quran_roots_{_roots_tag}_{_hash}"

# PDF
if "_export_pdf" in st.session_state and not st.session_state.get("_logged_pdf"):
    log_export("pdf"); st.session_state["_logged_pdf"] = True
if "_export_pdf" in st.session_state:
    st.download_button(
        f"PDF - {st.session_state.get('_export_pdf_count', 0)} pages "
        "(Reading Guide first, then every chart)",
        data=st.session_state["_export_pdf"],
        file_name=f"{_stem}.pdf",
        mime="application/pdf", width='stretch', type="primary")
else:
    st.info("PDF will appear here after you click **Build PDF + HTML** above.")

# HTML zip
if "_export_html_zip" in st.session_state and not st.session_state.get("_logged_html"):
    log_export("html"); st.session_state["_logged_html"] = True
if "_export_html_zip" in st.session_state:
    st.download_button(
        f"HTML zip - {st.session_state['_export_html_count']} files "
        "(Reading Guide + interactive charts)",
        data=st.session_state["_export_html_zip"],
        file_name=f"{_stem}_html.zip",
        mime="application/zip", width='stretch')
else:
    st.caption("Interactive HTML zip will appear here after you build.")

# Excel
if not st.session_state.get("_logged_xlsx_offer"):
    log_export("xlsx_offer"); st.session_state["_logged_xlsx_offer"] = True
st.download_button(
    "Excel workbook - 14 sheets (Reading Guide + raw tables)",
    data=_xlsx_bytes,
    file_name=f"{_stem}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    width='stretch')


st.divider()
layer(4, "Excel sheet contents")
st.markdown("""
- **Reading Guide** -- plain-English narrative (NEW)
- **Summary** -- meta info + per-root counts + motif overview
- **Occurrences** -- every (root, ayah) match row
- **Co-occurrence** -- every partner root + ayahs together
- **Surface Forms** -- distinct surface forms per input root
- **Partner Motifs** -- top partners per input root + affinity score
- **Triangles** -- every closed triad in the network, ranked by sum-weight
- **Centrality** -- degree / weighted-degree / betweenness / eigenvector / clustering
- **Surah Heatmap** -- root x surah ayah-hit matrix
- **Overlap Matrix** -- pairwise shared-ayah counts
- **Morphology** -- prefix/suffix particle counts per root
- **Position Stats** -- every match's position-in-ayah + ayah length
- **Baseline Rarity** -- corpus-baseline comparison (percentile, z-score, tier)
- **First & Last** -- first & last surah:ayah occurrence per root
""")
