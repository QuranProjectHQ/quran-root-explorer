"""Core analysis library — supports book5.xlsx and book6.xlsx (with diacritized col)."""
from __future__ import annotations

import io
import re
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from itertools import combinations
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import networkx as nx
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COL_SURAH = "ش  سوره"
COL_AYAH = "ش  آیه"
COL_SURAH_NAME = "اسم سوره"
COL_ROOTS = "ریشه نحوی"
COL_SURFACE = "توکن ریشه نحوی"
COL_SEGMENTED = "متن آیه توکن شده بی حرکت"
COL_DIACRITIZED = "متن آیه با حرکت"   # NEW: original Quranic text WITH diacritics (book6+)
COL_REV_ORDER  = "ترتیب نزول"          # NEW (book6 v2+): revelation order at surah scale (1..114)

# Egyptian-standard revelation-order cutoff (rev_order <= MECCAN_CUTOFF = Meccan)
MECCAN_CUTOFF = 86

_DIACRITICS = re.compile(r"[\u064B-\u065F\u0670\u06D6-\u06ED]")
_TATWEEL = re.compile(r"ـ")


def strip_diacritics(text):
    if not isinstance(text, str):
        return ""
    return _TATWEEL.sub("", _DIACRITICS.sub("", text)).strip()


def normalize_letters(text):
    """Fold every common Persian/Urdu/Sindhi/etc. variant of an Arabic
    letter to its canonical standard-Arabic codepoint, so the same root
    entered from any keyboard or copy-pasted from any source matches.

    Canonical = standard Arabic codepoints:
      alef:  آ (U+0622) · أ (U+0623) · إ (U+0625) · ٱ (U+0671) ·
             ٲ (U+0672) · ٳ (U+0673) · ا (U+0627)
             → all folded to bare ا (U+0627).
      yeh:   Persian ی (U+06CC) · alif-maqsura ى (U+0649) · ێ (U+06CE) ·
             ې (U+06D0) · ۍ (U+06CD) · ؠ (U+0620) · Arabic ي (U+064A)
             → all folded to Arabic ي.
      kaf:   Persian ک (U+06A9) · ڪ (U+06AA) · ګ (U+06AB) · Arabic ك (U+0643)
             → all folded to Arabic ك.
      heh:   ة (U+0629) · ھ (U+06BE) · ۀ (U+06C0) · ه (U+0647)
             → all folded to ه.
      waw:   ۆ (U+06C6) · ۇ (U+06C7) · ۈ (U+06C8) · ۉ (U+06C9) ·
             ۋ (U+06CB) · ۅ (U+06C5) · ۥ (U+06E5) · و (U+0648)
             → all folded to و.
      hamza: ؤ (U+0624) · ئ (U+0626) · ٔ (U+0654, hamza above) · ء (U+0621)
             → all folded to bare ء.
      Arabic-Indic digits ٠١٢٣٤٥٦٧٨٩ → 0123456789.
      Tatweel ـ stripped.  All diacritics stripped.
    """
    if not isinstance(text, str):
        return ""
    text = strip_diacritics(text)
    # alef variants → bare alef
    text = re.sub(r"[آأإٱٲٳ]", "ا", text)
    # yeh variants → Arabic ي
    text = re.sub(r"[یىێېۍؠ]", "ي", text)
    # ta-marbuta + Urdu heh variants → ه
    text = re.sub(r"[ةھۀ]", "ه", text)
    # kaf variants → Arabic ك
    text = re.sub(r"[کڪګ]", "ك", text)
    # waw variants → و
    text = re.sub(r"[ۆۇۈۉۋۅۥ]", "و", text)
    # hamza-bearing letters → bare hamza
    text = re.sub(r"[ؤئ]", "ء", text)
    # combining hamza above (U+0654) — strip
    text = text.replace("ٔ", "")
    # combining hamza below (U+0655) — strip
    text = text.replace("ٕ", "")
    # Arabic-Indic digits → ASCII
    digits = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
    text = text.translate(digits)
    return text.strip()


def tokenize(text):
    if not isinstance(text, str):
        return []
    return [t for t in text.split() if t]


@dataclass
class Corpus:
    df: pd.DataFrame
    root_tokens: list = field(default_factory=list)
    surface_tokens: list = field(default_factory=list)
    seg_tokens: list = field(default_factory=list)
    index_exact: dict = field(default_factory=dict)
    index_norm: dict = field(default_factory=dict)
    freq_exact: Counter = field(default_factory=Counter)
    freq_norm: Counter = field(default_factory=Counter)
    has_diacritized: bool = False
    has_rev_order: bool = False
    # When has_rev_order: surah->revelation_order lookup, and a parallel
    # global index (rev_global_idx, 0..N-1) for ayahs sorted in revelation order
    rev_order_of_surah: dict = field(default_factory=dict)
    rev_global_idx: list = field(default_factory=list)  # rev_global_idx[i] for df row i

    @property
    def n_ayahs(self): return len(self.df)
    @property
    def n_unique_roots(self): return len(self.index_exact)


def _detect_header_row(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    try:
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 25:
                break
            vals = [str(v) if v is not None else "" for v in row]
            if COL_SURAH in vals and COL_ROOTS in vals:
                return i
    finally:
        wb.close()
    return 11


def load_corpus(xlsx_path):
    header_row = _detect_header_row(xlsx_path)
    df = pd.read_excel(xlsx_path, header=header_row)
    base_cols = [COL_SURAH, COL_AYAH, COL_SURAH_NAME, COL_ROOTS, COL_SURFACE, COL_SEGMENTED]
    has_diacritized = COL_DIACRITIZED in df.columns
    has_rev_order = COL_REV_ORDER in df.columns
    cols = base_cols + ([COL_DIACRITIZED] if has_diacritized else []) \
                     + ([COL_REV_ORDER] if has_rev_order else [])
    df = df[cols].copy().reset_index(drop=True)
    df = df.dropna(subset=[COL_SURAH, COL_AYAH]).reset_index(drop=True)
    df[COL_ROOTS] = df[COL_ROOTS].fillna("").map(strip_diacritics)
    df[COL_SURFACE] = df[COL_SURFACE].fillna("").map(strip_diacritics)
    df[COL_SEGMENTED] = df[COL_SEGMENTED].fillna("").map(strip_diacritics)
    if has_diacritized:
        df[COL_DIACRITIZED] = df[COL_DIACRITIZED].fillna("").map(
            lambda s: s.strip() if isinstance(s, str) else "")
    if has_rev_order:
        df[COL_REV_ORDER] = pd.to_numeric(df[COL_REV_ORDER], errors="coerce").astype("Int64")

    root_tokens = [tokenize(s) for s in df[COL_ROOTS].tolist()]
    surface_tokens = [tokenize(s) for s in df[COL_SURFACE].tolist()]
    seg_tokens = [tokenize(s) for s in df[COL_SEGMENTED].tolist()]
    index_exact = defaultdict(list)
    index_norm = defaultdict(list)
    freq_exact = Counter()
    freq_norm = Counter()
    for i, toks in enumerate(root_tokens):
        seen_exact, seen_norm = set(), set()
        for t in toks:
            n = normalize_letters(t)
            if t not in seen_exact:
                index_exact[t].append(i); freq_exact[t] += 1; seen_exact.add(t)
            if n not in seen_norm:
                index_norm[n].append(i); freq_norm[n] += 1; seen_norm.add(n)

    # ---- Revelation-order indexing ----
    rev_order_of_surah = {}
    rev_global_idx = [0] * len(df)
    if has_rev_order:
        # Build surah->rev_order lookup from the data (most common value per surah)
        for s, sub in df.groupby(COL_SURAH):
            vals = sub[COL_REV_ORDER].dropna().tolist()
            if vals:
                rev_order_of_surah[int(s)] = int(vals[0])
        # Compute rev_global_idx: sort by (rev_order, ayah_within_surah), assign 0..N-1
        keyed = [(rev_order_of_surah.get(int(df.iloc[i][COL_SURAH]), 999),
                  int(df.iloc[i][COL_AYAH]), i) for i in range(len(df))]
        keyed.sort()
        for new_idx, (_, _, orig_i) in enumerate(keyed):
            rev_global_idx[orig_i] = new_idx

    return Corpus(df=df, root_tokens=root_tokens, surface_tokens=surface_tokens,
                  seg_tokens=seg_tokens, index_exact=dict(index_exact),
                  index_norm=dict(index_norm), freq_exact=freq_exact,
                  freq_norm=freq_norm, has_diacritized=has_diacritized,
                  has_rev_order=has_rev_order,
                  rev_order_of_surah=rev_order_of_surah,
                  rev_global_idx=rev_global_idx)


def parse_input_roots(text, normalize):
    toks = tokenize(strip_diacritics(text or ""))
    if normalize:
        toks = [normalize_letters(t) for t in toks]
    seen, out = set(), []
    for t in toks:
        if t and t not in seen:
            seen.add(t); out.append(t)
    return out


def search_root(corpus, root, normalize):
    if normalize:
        return corpus.index_norm.get(normalize_letters(root), [])
    return corpus.index_exact.get(root, [])


def find_occurrences(corpus, input_roots, normalize):
    rows = []
    for q in input_roots:
        for i in search_root(corpus, q, normalize):
            r = corpus.df.iloc[i]
            r_tokens = corpus.root_tokens[i]
            s_tokens = corpus.surface_tokens[i]
            positions = [j for j, t in enumerate(r_tokens)
                         if (normalize_letters(t) if normalize else t) == q]
            matched_surface = [s_tokens[j] if j < len(s_tokens) else "" for j in positions]
            row_dict = {
                "Input Root": q,
                "Surah #": int(r[COL_SURAH]),
                "Ayah #": int(r[COL_AYAH]),
                "Surah Name": r[COL_SURAH_NAME],
                "Surface Form(s)": " ، ".join(matched_surface),
                "All Roots in Ayah": r[COL_ROOTS],
                "All Surface Tokens": r[COL_SURFACE],
                "Segmented Ayah": r[COL_SEGMENTED],
                "Hit Count": len(positions),
            }
            if corpus.has_diacritized:
                row_dict["Quranic Text (diacritized)"] = r[COL_DIACRITIZED]
            rows.append(row_dict)
    if not rows:
        cols = ["Input Root", "Surah #", "Ayah #", "Surah Name", "Surface Form(s)",
                "All Roots in Ayah", "All Surface Tokens", "Segmented Ayah", "Hit Count"]
        if corpus.has_diacritized:
            cols.append("Quranic Text (diacritized)")
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(rows).sort_values(["Input Root", "Surah #", "Ayah #"]).reset_index(drop=True)


def cooccurrence(corpus, input_roots, normalize):
    match_ayahs = set()
    for q in input_roots:
        match_ayahs.update(search_root(corpus, q, normalize))
    input_set = set(input_roots)
    K = (normalize_letters if normalize else (lambda t: t))
    partners = Counter()
    for i in match_ayahs:
        for k in {K(t) for t in corpus.root_tokens[i]} - input_set:
            partners[k] += 1
    return partners, match_ayahs


def cooccurrence_table(partners, top_n=None):
    items = partners.most_common(top_n) if top_n else partners.most_common()
    return pd.DataFrame(items, columns=["Co-occurring Root", "Ayahs Together"])


def build_network(corpus, input_roots, normalize, top_partners=15, min_weight=1):
    partners, match_ayahs = cooccurrence(corpus, input_roots, normalize)
    keep = {p for p, _ in partners.most_common(top_partners)}
    node_set = set(input_roots) | keep
    K = (normalize_letters if normalize else (lambda t: t))
    edge_w = Counter()
    for i in match_ayahs:
        roots_here = list({K(t) for t in corpus.root_tokens[i]} & node_set)
        for a, b in combinations(sorted(roots_here), 2):
            edge_w[(a, b)] += 1
    g = nx.Graph()
    for n in node_set:
        g.add_node(n, is_input=(n in set(input_roots)))
    for (a, b), w in edge_w.items():
        if w >= min_weight:
            g.add_edge(a, b, weight=w)
    return g


def triad_census(g):
    triangles = open_triads = 0
    for a, b, c in combinations(g.nodes(), 3):
        deg = g.has_edge(a, b) + g.has_edge(b, c) + g.has_edge(a, c)
        if deg == 3: triangles += 1
        elif deg == 2: open_triads += 1
    return {
        "triangles (closed triads)": triangles,
        "open triads (paths of length 2)": open_triads,
        "edges": g.number_of_edges(),
        "nodes": g.number_of_nodes(),
        "density": round(nx.density(g), 4) if g.number_of_nodes() > 1 else 0.0,
    }


def triangles_table(g, limit=100):
    rows, seen = [], set()
    for a, b, c in combinations(g.nodes(), 3):
        if g.has_edge(a, b) and g.has_edge(b, c) and g.has_edge(a, c):
            key = tuple(sorted((a, b, c)))
            if key in seen: continue
            seen.add(key)
            w = g[a][b]["weight"] + g[b][c]["weight"] + g[a][c]["weight"]
            inputs_in = sum(1 for n in key if g.nodes[n].get("is_input"))
            rows.append({"Root A": key[0], "Root B": key[1], "Root C": key[2],
                         "Sum Weight": w, "Inputs in Triad": inputs_in})
    if not rows:
        return pd.DataFrame(columns=["Root A", "Root B", "Root C", "Sum Weight", "Inputs in Triad"])
    return pd.DataFrame(rows).sort_values("Sum Weight", ascending=False).head(limit).reset_index(drop=True)


def partner_motifs(corpus, input_roots, normalize, top=20):
    K = (normalize_letters if normalize else (lambda t: t))
    rows = []
    for q in input_roots:
        ayahs = search_root(corpus, q, normalize)
        c = Counter()
        for i in ayahs:
            for k in {K(t) for t in corpus.root_tokens[i]} - {q}:
                c[k] += 1
        for p, n in c.most_common(top):
            rows.append({"Input Root": q, "Partner Root": p, "Ayahs Together": n,
                         "Total Ayahs of Input": len(ayahs),
                         "Affinity": round(n / max(len(ayahs), 1), 4)})
    if not rows:
        return pd.DataFrame(columns=["Input Root", "Partner Root", "Ayahs Together",
                                     "Total Ayahs of Input", "Affinity"])
    return pd.DataFrame(rows)


def triangle_ayahs(corpus, triad, normalize):
    sets = [set(search_root(corpus, t, normalize)) for t in triad]
    if not sets:
        return pd.DataFrame()
    common = sets[0]
    for s in sets[1:]: common &= s
    rows = []
    for i in sorted(common):
        r = corpus.df.iloc[i]
        row = {"Surah #": int(r[COL_SURAH]), "Ayah #": int(r[COL_AYAH]),
               "Surah Name": r[COL_SURAH_NAME], "All Roots": r[COL_ROOTS],
               "Segmented Ayah": r[COL_SEGMENTED]}
        if corpus.has_diacritized:
            row["Quranic Text (diacritized)"] = r[COL_DIACRITIZED]
        rows.append(row)
    return pd.DataFrame(rows)


def surface_form_table(corpus, input_roots, normalize):
    K = (normalize_letters if normalize else (lambda t: t))
    rows = []
    for q in input_roots:
        forms = Counter()
        contexts_seg = {}
        contexts_dia = {}
        for i in search_root(corpus, q, normalize):
            r_tokens = corpus.root_tokens[i]
            s_tokens = corpus.surface_tokens[i]
            r = corpus.df.iloc[i]
            for j, t in enumerate(r_tokens):
                if K(t) == q and j < len(s_tokens):
                    sf = s_tokens[j]
                    forms[sf] += 1
                    contexts_seg.setdefault(sf, r[COL_SEGMENTED])
                    if corpus.has_diacritized:
                        contexts_dia.setdefault(sf, r[COL_DIACRITIZED])
        for sf, n in forms.most_common():
            d = {"Input Root": q, "Surface Form (col 5)": sf, "Occurrences": n,
                 "Example Segmented (col 6)": contexts_seg.get(sf, "")}
            if corpus.has_diacritized:
                d["Example Diacritized (col 7)"] = contexts_dia.get(sf, "")
            rows.append(d)
    if not rows:
        cols = ["Input Root", "Surface Form (col 5)", "Occurrences", "Example Segmented (col 6)"]
        if corpus.has_diacritized:
            cols.append("Example Diacritized (col 7)")
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(rows)


def surah_heatmap(corpus, input_roots, normalize):
    if not input_roots:
        return pd.DataFrame()
    data = {q: defaultdict(int) for q in input_roots}
    for q in input_roots:
        for i in search_root(corpus, q, normalize):
            data[q][int(corpus.df.iloc[i][COL_SURAH])] += 1
    all_surahs = sorted({s for d in data.values() for s in d.keys()}) or list(range(1, 115))
    df = pd.DataFrame(0, index=input_roots, columns=all_surahs, dtype=int)
    for q in input_roots:
        for s, v in data[q].items():
            df.loc[q, s] = v
    df.index.name = "Root"; df.columns.name = "Surah #"
    return df


def overlap_matrix(corpus, input_roots, normalize):
    ayahs = {q: set(search_root(corpus, q, normalize)) for q in input_roots}
    m = pd.DataFrame(0, index=input_roots, columns=input_roots, dtype=int)
    for a in input_roots:
        for b in input_roots:
            m.loc[a, b] = len(ayahs[a] & ayahs[b])
    return m


def centrality_table(g):
    if g.number_of_nodes() == 0:
        return pd.DataFrame(columns=["Root", "Degree", "Weighted Degree", "Betweenness",
                                     "Eigenvector", "Clustering", "Is Input"])
    deg = dict(g.degree())
    wdeg = dict(g.degree(weight="weight"))
    try: btw = nx.betweenness_centrality(g, weight="weight", normalized=True)
    except Exception: btw = {n: 0.0 for n in g.nodes()}
    try: eig = nx.eigenvector_centrality_numpy(g, weight="weight")
    except Exception: eig = {n: 0.0 for n in g.nodes()}
    clust = nx.clustering(g, weight="weight")
    rows = [{
        "Root": n, "Degree": deg.get(n, 0),
        "Weighted Degree": round(wdeg.get(n, 0), 2),
        "Betweenness": round(btw.get(n, 0), 4),
        "Eigenvector": round(eig.get(n, 0), 4),
        "Clustering": round(clust.get(n, 0), 4),
        "Is Input": g.nodes[n].get("is_input", False),
    } for n in g.nodes()]
    return pd.DataFrame(rows).sort_values("Weighted Degree", ascending=False).reset_index(drop=True)


def detect_communities(g):
    if g.number_of_nodes() == 0 or g.number_of_edges() == 0:
        return {n: 0 for n in g.nodes()}
    try:
        from networkx.algorithms.community import greedy_modularity_communities
        comms = list(greedy_modularity_communities(g, weight="weight"))
        return {n: idx for idx, c in enumerate(comms) for n in c}
    except Exception:
        return {n: 0 for n in g.nodes()}


_ARABIC_PARTICLES = {
    "ال": "definite article 'al-'", "و": "and (wa)", "ف": "then/so (fa)",
    "ب": "by/with (bi)", "ل": "for/to (li)", "ك": "like/as (ka)",
    "س": "future marker (sa)", "ه": "his/him", "ها": "her",
    "هم": "their/them (m.)", "هن": "their/them (f.)", "نا": "our/us",
    "كم": "your (pl.)", "ني": "me", "ي": "my",
}


def morphology_breakdown(corpus, input_roots, normalize):
    K = (normalize_letters if normalize else (lambda t: t))
    rows = []
    for q in input_roots:
        pre, suf = Counter(), Counter()
        for i in search_root(corpus, q, normalize):
            r_tokens = corpus.root_tokens[i]
            s_tokens = corpus.surface_tokens[i]
            seg_tokens = corpus.seg_tokens[i]
            targets = [j for j, r in enumerate(r_tokens) if K(r) == q]
            for pos in targets:
                if pos >= len(s_tokens): continue
                sf = s_tokens[pos]
                for sidx, segtok in enumerate(seg_tokens):
                    if (segtok == sf or sf.startswith(segtok)
                            or (len(sf) >= 3 and segtok.startswith(sf[:max(3, len(sf) // 2)]))):
                        for d in (1, 2):
                            if sidx - d >= 0 and seg_tokens[sidx - d] in _ARABIC_PARTICLES:
                                pre[seg_tokens[sidx - d]] += 1
                            if sidx + d < len(seg_tokens) and seg_tokens[sidx + d] in _ARABIC_PARTICLES:
                                suf[seg_tokens[sidx + d]] += 1
                        break
        for p, c in pre.most_common():
            rows.append({"Input Root": q, "Particle": p, "Position": "prefix",
                         "Meaning": _ARABIC_PARTICLES.get(p, ""), "Count": c})
        for p, c in suf.most_common():
            rows.append({"Input Root": q, "Particle": p, "Position": "suffix",
                         "Meaning": _ARABIC_PARTICLES.get(p, ""), "Count": c})
    if not rows:
        return pd.DataFrame(columns=["Input Root", "Particle", "Position", "Meaning", "Count"])
    return pd.DataFrame(rows)


def position_stats(corpus, input_roots, normalize):
    K = (normalize_letters if normalize else (lambda t: t))
    rows = []
    for q in input_roots:
        for i in search_root(corpus, q, normalize):
            r_tokens = corpus.root_tokens[i]
            n = len(r_tokens)
            for j, t in enumerate(r_tokens):
                if K(t) == q:
                    rel = round(j / max(n - 1, 1), 4) if n > 1 else 0.5
                    rows.append({
                        "Input Root": q, "Surah #": int(corpus.df.iloc[i][COL_SURAH]),
                        "Ayah #": int(corpus.df.iloc[i][COL_AYAH]),
                        "Position in ayah (0..1)": rel, "Token index": j,
                        "Ayah length (roots)": n,
                    })
    if not rows:
        return pd.DataFrame(columns=["Input Root", "Surah #", "Ayah #",
                                     "Position in ayah (0..1)", "Token index",
                                     "Ayah length (roots)"])
    return pd.DataFrame(rows)


def baseline_rarity(corpus, input_roots, normalize):
    freqs = corpus.freq_norm if normalize else corpus.freq_exact
    counts = sorted(freqs.values())
    median = counts[len(counts) // 2] if counts else 0
    mean = statistics.mean(counts) if counts else 0
    stdev = statistics.pstdev(counts) if counts else 1
    rows = []
    for q in input_roots:
        c = freqs.get(q, 0)
        rank = sum(1 for v in counts if v < c)
        pct = round(100 * rank / max(len(counts), 1), 2)
        z = round((c - mean) / stdev, 2) if stdev else 0
        tier = ("ultra-rare" if pct < 10 else "rare" if pct < 33
                else "common" if pct < 66 else "very common" if pct < 90 else "ubiquitous")
        rows.append({"Input Root": q, "Ayah Frequency": c, "Corpus Median": median,
                     "Corpus Mean": round(mean, 2), "Percentile": pct,
                     "Z-score": z, "Tier": tier})
    return pd.DataFrame(rows)


def first_last_occurrence(corpus, input_roots, normalize):
    rows = []
    for q in input_roots:
        idx = search_root(corpus, q, normalize)
        if not idx:
            rows.append({"Input Root": q, "First (S:A)": "—", "First Surah Name": "—",
                         "Last (S:A)": "—", "Last Surah Name": "—"})
            continue
        first = corpus.df.iloc[idx[0]]
        last = corpus.df.iloc[idx[-1]]
        rows.append({
            "Input Root": q,
            "First (S:A)": f"{int(first[COL_SURAH])}:{int(first[COL_AYAH])}",
            "First Surah Name": first[COL_SURAH_NAME],
            "Last (S:A)": f"{int(last[COL_SURAH])}:{int(last[COL_AYAH])}",
            "Last Surah Name": last[COL_SURAH_NAME],
        })
    return pd.DataFrame(rows)


def summary_stats(corpus, input_roots, occurrences, partners):
    rows = []
    for q in input_roots:
        sub = occurrences[occurrences["Input Root"] == q]
        n_ayahs = sub[["Surah #", "Ayah #"]].drop_duplicates().shape[0]
        n_surahs = sub["Surah #"].nunique() if not sub.empty else 0
        total_hits = int(sub["Hit Count"].sum()) if not sub.empty else 0
        rows.append({"Input Root": q, "Ayahs Found": n_ayahs,
                     "Surahs Covered": n_surahs, "Total Hits": total_hits})
    rows.append({
        "Input Root": "— ALL INPUT ROOTS (any) —",
        "Ayahs Found": occurrences[["Surah #", "Ayah #"]].drop_duplicates().shape[0] if not occurrences.empty else 0,
        "Surahs Covered": occurrences["Surah #"].nunique() if not occurrences.empty else 0,
        "Total Hits": int(occurrences["Hit Count"].sum()) if not occurrences.empty else 0,
    })
    rows.append({"Input Root": "— Unique partner roots —", "Ayahs Found": "",
                 "Surahs Covered": "", "Total Hits": len(partners)})
    return pd.DataFrame(rows)


# Matplotlib export plots
def _shape_ar(text):
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        return get_display(arabic_reshaper.reshape(text))
    except Exception:
        return text


def plot_network(g, input_roots, title="Root Co-occurrence Network"):
    fig, ax = plt.subplots(figsize=(11, 8.5))
    if g.number_of_nodes() == 0:
        ax.text(0.5, 0.5, "No matches", ha="center"); ax.axis("off"); return fig
    pos = nx.spring_layout(g, seed=42, k=1.4 / max(g.number_of_nodes(), 1) ** 0.5)
    node_colors = ["#E63946" if g.nodes[n].get("is_input") else "#457B9D" for n in g.nodes()]
    node_sizes = [1500 if g.nodes[n].get("is_input") else 800 for n in g.nodes()]
    ew = [g[u][v]["weight"] for u, v in g.edges()]
    mw = max(ew) if ew else 1
    widths = [0.5 + 4 * (w / mw) for w in ew]
    nx.draw_networkx_edges(g, pos, width=widths, edge_color="#888", alpha=0.5, ax=ax)
    nx.draw_networkx_nodes(g, pos, node_color=node_colors, node_size=node_sizes,
                           edgecolors="black", linewidths=1.2, ax=ax)
    labels = {n: _shape_ar(n) for n in g.nodes()}
    nx.draw_networkx_labels(g, pos, labels=labels, font_size=11, font_family="DejaVu Sans", ax=ax)
    ax.set_title(title, fontsize=14, pad=12); ax.axis("off")
    fig.tight_layout()
    return fig


def plot_top_partners(partners, top=20, title="Top Co-occurring Roots"):
    fig, ax = plt.subplots(figsize=(10, 7))
    items = partners.most_common(top)
    if not items:
        ax.text(0.5, 0.5, "None", ha="center"); ax.axis("off"); return fig
    labels = [_shape_ar(k) for k, _ in items][::-1]
    values = [v for _, v in items][::-1]
    ax.barh(labels, values, color="#457B9D", edgecolor="black")
    ax.set_xlabel("Co-occurring ayah count")
    ax.set_title(title, fontsize=14, pad=10)
    for i, v in enumerate(values):
        ax.text(v, i, f" {v}", va="center", fontsize=9)
    fig.tight_layout()
    return fig


def plot_surah_distribution(occurrences, title="Distribution Across Surahs"):
    fig, ax = plt.subplots(figsize=(12, 5))
    if occurrences.empty:
        ax.text(0.5, 0.5, "No matches", ha="center"); ax.axis("off"); return fig
    g = occurrences.groupby(["Surah #", "Input Root"]).size().unstack(fill_value=0).sort_index()
    g.plot(kind="bar", stacked=True, ax=ax, width=0.9, colormap="viridis")
    ax.set_xlabel("Surah #"); ax.set_ylabel("Ayah hits")
    ax.set_title(title, fontsize=14, pad=10)
    h, l = ax.get_legend_handles_labels()
    ax.legend(h, [_shape_ar(x) for x in l], title="Input root", loc="upper right")
    fig.tight_layout()
    return fig


def plot_triad_summary(triad, title="Motif Summary"):
    fig, ax = plt.subplots(figsize=(8, 5))
    keys = ["nodes", "edges", "triangles (closed triads)", "open triads (paths of length 2)"]
    values = [triad.get(k, 0) for k in keys]
    ax.bar(keys, values, color=["#888", "#aaa", "#E63946", "#457B9D"], edgecolor="black")
    ax.set_title(title, fontsize=14, pad=10)
    for i, v in enumerate(values):
        ax.text(i, v, f" {v}", ha="center", va="bottom", fontsize=10)
    plt.xticks(rotation=20, ha="right")
    fig.tight_layout()
    return fig


def figures_to_pdf(figures, pdf_path):
    from matplotlib.backends.backend_pdf import PdfPages
    with PdfPages(pdf_path) as pdf:
        for f in figures:
            pdf.savefig(f, bbox_inches="tight")
    return Path(pdf_path)


def figure_to_png_bytes(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    return buf.getvalue()


# Excel export
HEADER_FILL = PatternFill("solid", start_color="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
BODY_FONT = Font(name="Arial")


def _style_sheet(ws, df):
    for col_idx, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=str(col))
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df.iloc[:, col_idx - 1].head(200).tolist()])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)
    ws.freeze_panes = "A2"


def export_excel(out_path, summary, occurrences, cooccurrence_tbl, surface_forms,
                 partner_motifs_tbl, triangles_tbl, triad_summary, meta,
                 centrality=None, communities=None, heatmap=None, overlap=None,
                 morphology=None, position=None, rarity=None, first_last=None):
    wb = Workbook()
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = "Quran Root Analysis — Summary"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws.merge_cells("A1:D1")
    row = 3
    for k, v in meta.items():
        ws.cell(row=row, column=1, value=k).font = Font(name="Arial", bold=True)
        ws.cell(row=row, column=2, value=str(v)).font = BODY_FONT
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Per-root summary").font = Font(name="Arial", bold=True, size=12)
    row += 1
    for col_idx, col in enumerate(summary.columns, start=1):
        c = ws.cell(row=row, column=col_idx, value=str(col))
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    for r_off, rec in enumerate(summary.itertuples(index=False), start=1):
        for col_idx, val in enumerate(rec, start=1):
            ws.cell(row=row + r_off, column=col_idx, value=val).font = BODY_FONT
    for col_idx in range(1, len(summary.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 26
    base = row + len(summary) + 3
    ws.cell(row=base, column=1, value="Motif / network summary").font = Font(name="Arial", bold=True, size=12)
    for i, (k, v) in enumerate(triad_summary.items(), start=1):
        ws.cell(row=base + i, column=1, value=k).font = Font(name="Arial", bold=True)
        ws.cell(row=base + i, column=2, value=v).font = BODY_FONT

    sheets = [
        ("Occurrences", occurrences),
        ("Co-occurrence", cooccurrence_tbl),
        ("Surface Forms", surface_forms),
        ("Partner Motifs", partner_motifs_tbl),
        ("Triangles", triangles_tbl),
    ]
    if centrality is not None: sheets.append(("Centrality", centrality))
    if heatmap is not None: sheets.append(("Surah Heatmap", heatmap.reset_index()))
    if overlap is not None: sheets.append(("Overlap Matrix", overlap.reset_index()))
    if morphology is not None: sheets.append(("Morphology", morphology))
    if position is not None: sheets.append(("Position Stats", position))
    if rarity is not None: sheets.append(("Baseline Rarity", rarity))
    if first_last is not None: sheets.append(("First & Last", first_last))

    for name, df in sheets:
        sheet = wb.create_sheet(name)
        if df is None or df.empty:
            sheet["A1"] = f"(no data for {name})"
        else:
            _style_sheet(sheet, df)
    wb.save(out_path)
    return Path(out_path)


# ===========================================================================
# ENRICHED NETWORK ATTRIBUTES — positional, spatial, rhythm, lead-lag
# ===========================================================================
# These functions exploit the structural information we have per occurrence:
#   - global ayah index i (position in mushaf, 0..6235)
#   - surah number, ayah-within-surah
#   - token index within ayah, ayah length in tokens
# and turn it into node/edge attributes plus several new tables.
#
# CAVEATS (disclose to the user in the UI, not silently):
#   * "Position in mushaf" is structural ordering, NOT historical/revelation
#     order — book6 carries no revelation column.
#   * Token positions reflect the corpus's tokenization choices.


def _root_positions(corpus, input_roots, normalize, order="mushaf"):
    """For each root, return (global_idx, surah, ayah, token_idx, ayah_len,
    mushaf_idx, rev_idx) tuples. The leading `global_idx` is mushaf_idx when
    order='mushaf' and rev_idx when order='revelation', so callers that read
    items[0] for the X-axis automatically get the right ordering. The
    explicit mushaf_idx and rev_idx are kept too for any caller that needs
    both."""
    K = (normalize_letters if normalize else (lambda t: t))
    out = {r: [] for r in input_roots}
    targets = {r: K(r) for r in input_roots}
    use_rev = (order == "revelation" and corpus.has_rev_order)
    for i in range(len(corpus.df)):
        toks = corpus.root_tokens[i]
        n = len(toks)
        if n == 0:
            continue
        s_num = int(corpus.df.iloc[i][COL_SURAH])
        a_num = int(corpus.df.iloc[i][COL_AYAH])
        mushaf_idx = i
        rev_idx = corpus.rev_global_idx[i] if use_rev or corpus.has_rev_order else i
        gidx = rev_idx if use_rev else mushaf_idx
        for j, t in enumerate(toks):
            kt = K(t)
            for r, rk in targets.items():
                if kt == rk:
                    out[r].append((gidx, s_num, a_num, j, n,
                                   mushaf_idx, rev_idx))
    for r in out:
        out[r].sort()
    return out


def _entropy(counts):
    """Shannon entropy in bits."""
    import math
    total = sum(counts)
    if total <= 0:
        return 0.0
    h = 0.0
    for c in counts:
        if c <= 0:
            continue
        p = c / total
        h -= p * math.log2(p)
    return h


def _fano_factor(gaps):
    """variance / mean of gaps. >1 = bursty, <1 = regular."""
    if not gaps:
        return 0.0
    mean = sum(gaps) / len(gaps)
    if mean == 0:
        return 0.0
    var = sum((g - mean) ** 2 for g in gaps) / len(gaps)
    return var / mean


def node_attributes(corpus, input_roots, normalize, order="mushaf"):
    """Per-root structural attributes — feeds enriched topology, spatial
    view, fingerprint radar, and node coloring.

    `order` controls how gravitational center, burstiness, and first/last
    indices are computed: 'mushaf' (canonical reading order) or 'revelation'
    (Egyptian-standard revelation order, surah-scale).

    When revelation-order data is present we also add Meccan / Medinan
    ayah-count columns (rev_order <= 86 = Meccan, otherwise Medinan).
    """
    import pandas as _pd
    pos = _root_positions(corpus, input_roots, normalize, order=order)
    rows = []
    surah_lengths = (corpus.df.groupby(COL_SURAH)[COL_AYAH].max().astype(int)
                     .to_dict())
    has_rev = corpus.has_rev_order
    rev_lookup = corpus.rev_order_of_surah

    def _empty():
        d = {
            "Root": "", "Total": 0, "Surahs Covered": 0,
            "Gravitational Center": 0, "Spread (Entropy)": 0.0,
            "Burstiness (Fano)": 0.0, "Mean Pos-in-Ayah": 0.0,
            "Mean Pos-in-Surah": 0.0, "Peak Surah": 0,
            "First Surah": 0, "Last Surah": 0,
            "First Global Idx": 0, "Last Global Idx": 0,
        }
        if has_rev:
            d["First Rev-Order Surah"] = 0
            d["Meccan Ayahs"] = 0
            d["Medinan Ayahs"] = 0
            d["Meccan %"] = 0.0
        return d

    for r in input_roots:
        items = pos[r]
        if not items:
            row = _empty(); row["Root"] = r
            rows.append(row); continue
        global_idxs = [it[0] for it in items]
        surahs = [it[1] for it in items]
        token_idxs = [it[3] for it in items]
        ayah_lens = [it[4] for it in items]
        grav = sum(global_idxs) / len(global_idxs)
        s_counts = Counter(surahs)
        spread = _entropy(s_counts.values())
        gaps = [global_idxs[k + 1] - global_idxs[k]
                for k in range(len(global_idxs) - 1)]
        burst = _fano_factor(gaps)
        pos_in_ayah_vals = [tj / max(al - 1, 1) if al > 1 else 0.5
                            for tj, al in zip(token_idxs, ayah_lens)]
        mean_pos_ayah = sum(pos_in_ayah_vals) / len(pos_in_ayah_vals)
        pos_in_surah_vals = []
        for it in items:
            s = it[1]; a = it[2]
            sl = surah_lengths.get(s, a)
            pos_in_surah_vals.append(a / max(sl, 1))
        mean_pos_surah = sum(pos_in_surah_vals) / len(pos_in_surah_vals)
        peak_surah = s_counts.most_common(1)[0][0]
        row = {
            "Root": r,
            "Total": len(global_idxs),
            "Surahs Covered": len(s_counts),
            "Gravitational Center": round(grav, 1),
            "Spread (Entropy)": round(spread, 3),
            "Burstiness (Fano)": round(burst, 3),
            "Mean Pos-in-Ayah": round(mean_pos_ayah, 3),
            "Mean Pos-in-Surah": round(mean_pos_surah, 3),
            "Peak Surah": peak_surah,
            "First Surah": surahs[0],
            "Last Surah": surahs[-1],
            "First Global Idx": global_idxs[0],
            "Last Global Idx": global_idxs[-1],
        }
        if has_rev:
            # Distinct ayahs per Meccan/Medinan (count unique mushaf indices)
            seen_idx = set()
            meccan = medinan = 0
            for it in items:
                mu = it[5]
                if mu in seen_idx:
                    continue
                seen_idx.add(mu)
                ro = rev_lookup.get(it[1], 999)
                if ro <= MECCAN_CUTOFF:
                    meccan += 1
                else:
                    medinan += 1
            total_unique = meccan + medinan
            row["First Rev-Order Surah"] = min(rev_lookup.get(s, 999)
                                                for s in s_counts) if s_counts else 0
            row["Meccan Ayahs"] = meccan
            row["Medinan Ayahs"] = medinan
            row["Meccan %"] = round(100 * meccan / max(total_unique, 1), 1)
        rows.append(row)
    return _pd.DataFrame(rows)


def edge_attributes(corpus, g, normalize):
    """Per-edge enrichment for the graph: mean token-distance, lead-lag,
    joint surahs, median ayah-gap, joint burstiness."""
    import pandas as _pd
    K = (normalize_letters if normalize else (lambda t: t))
    rows = []
    node_ayahs = {n: set(search_root(corpus, n, normalize)) for n in g.nodes()}
    for u, v, d in g.edges(data=True):
        joint = sorted(node_ayahs[u] & node_ayahs[v])
        if not joint:
            rows.append({"Root A": u, "Root B": v,
                         "Weight (ayahs)": d.get("weight", 0),
                         "Mean Token-Distance": 0.0,
                         "Lead-Lag (A->B)": 0.0,
                         "Joint Surahs": 0,
                         "Median Joint Gap": 0,
                         "Joint Burstiness": 0.0})
            continue
        distances = []
        a_leads = b_leads = 0
        surahs = set()
        ku, kv = K(u), K(v)
        for i in joint:
            toks = corpus.root_tokens[i]
            pa = [j for j, t in enumerate(toks) if K(t) == ku]
            pb = [j for j, t in enumerate(toks) if K(t) == kv]
            if not pa or not pb:
                continue
            for ja in pa:
                for jb in pb:
                    distances.append(abs(ja - jb))
            if min(pa) < min(pb):
                a_leads += 1
            elif min(pb) < min(pa):
                b_leads += 1
            surahs.add(int(corpus.df.iloc[i][COL_SURAH]))
        total_ll = a_leads + b_leads
        ll_score = ((a_leads - b_leads) / total_ll) if total_ll else 0.0
        mean_td = (sum(distances) / len(distances)) if distances else 0.0
        gaps = [joint[k + 1] - joint[k] for k in range(len(joint) - 1)]
        median_gap = (sorted(gaps)[len(gaps) // 2] if gaps else 0)
        rows.append({
            "Root A": u, "Root B": v,
            "Weight (ayahs)": d.get("weight", 0),
            "Mean Token-Distance": round(mean_td, 2),
            "Lead-Lag (A->B)": round(ll_score, 3),
            "Joint Surahs": len(surahs),
            "Median Joint Gap": median_gap,
            "Joint Burstiness": round(_fano_factor(gaps), 3),
        })
    if not rows:
        return _pd.DataFrame(columns=["Root A", "Root B", "Weight (ayahs)",
                                      "Mean Token-Distance", "Lead-Lag (A->B)",
                                      "Joint Surahs", "Median Joint Gap",
                                      "Joint Burstiness"])
    return _pd.DataFrame(rows).sort_values(
        "Weight (ayahs)", ascending=False).reset_index(drop=True)


def spatial_occurrences(corpus, input_roots, normalize, order="mushaf"):
    """Long-form table of every occurrence. Includes both Mushaf Idx and
    (when revelation-order data is available) Rev Idx columns plus a
    Meccan/Medinan label."""
    import pandas as _pd
    pos = _root_positions(corpus, input_roots, normalize, order=order)
    surah_lengths = (corpus.df.groupby(COL_SURAH)[COL_AYAH].max().astype(int)
                     .to_dict())
    has_rev = corpus.has_rev_order
    rev_lookup = corpus.rev_order_of_surah
    rows = []
    for r, items in pos.items():
        for (gidx, s, a, tj, al, mu_idx, rv_idx) in items:
            sl = surah_lengths.get(s, a)
            row = {
                "Root": r, "Surah": s, "Ayah": a,
                "Mushaf Idx": mu_idx,
                "Pos-in-Ayah": round(tj / max(al - 1, 1) if al > 1 else 0.5, 3),
                "Pos-in-Surah": round(a / max(sl, 1), 3),
                "Token Index": tj, "Ayah Length": al,
            }
            if has_rev:
                ro = rev_lookup.get(s, 999)
                row["Rev Idx"] = rv_idx
                row["Rev Order (surah)"] = ro
                row["Phase"] = "Meccan" if ro <= MECCAN_CUTOFF else "Medinan"
            rows.append(row)
    cols = ["Root", "Surah", "Ayah", "Mushaf Idx",
            "Pos-in-Ayah", "Pos-in-Surah", "Token Index", "Ayah Length"]
    if has_rev:
        cols += ["Rev Idx", "Rev Order (surah)", "Phase"]
    if not rows:
        return _pd.DataFrame(columns=cols)
    return _pd.DataFrame(rows)[cols]


def cumulative_trajectories(corpus, input_roots, normalize, order="mushaf"):
    """For each root, every global-index at which it occurs in the chosen
    order, plus running count. Drives the Rhythm & Growth chart."""
    import pandas as _pd
    pos = _root_positions(corpus, input_roots, normalize, order=order)
    rows = []
    for r, items in pos.items():
        seen_keys = set()
        running = 0
        for (gidx, s, a, tj, al, mu, rv) in items:
            if gidx in seen_keys:
                continue
            seen_keys.add(gidx)
            running += 1
            rows.append({"Root": r, "Global Idx": gidx,
                         "Surah": s, "Ayah": a,
                         "Cumulative Count": running})
    cols = ["Root", "Global Idx", "Surah", "Ayah", "Cumulative Count"]
    if not rows:
        return _pd.DataFrame(columns=cols)
    return _pd.DataFrame(rows)[cols]


def lead_lag_matrix(corpus, input_roots, normalize, window=2):
    """Directed P(B appears within +/-window MUSHAF ayahs of A | A appears).
    Lead-lag is intentionally mushaf-only: it measures textual adjacency."""
    import pandas as _pd
    pos = _root_positions(corpus, input_roots, normalize, order="mushaf")
    ayahs_of = {r: set(it[0] for it in items) for r, items in pos.items()}
    m = _pd.DataFrame(0.0, index=input_roots, columns=input_roots)
    for a in input_roots:
        ayahs_a = ayahs_of[a]
        if not ayahs_a:
            continue
        for b in input_roots:
            ayahs_b = ayahs_of[b]
            if a == b:
                m.loc[a, b] = 1.0
                continue
            hits = 0
            for i in ayahs_a:
                for k in range(-window, window + 1):
                    if (i + k) in ayahs_b:
                        hits += 1
                        break
            m.loc[a, b] = round(hits / len(ayahs_a), 3)
    return m


def fingerprint_table(corpus, input_roots, normalize, node_attrs=None,
                     order="mushaf"):
    """Normalized 0..1 axes per root for the radar/fingerprint chart."""
    import math
    import pandas as _pd
    if node_attrs is None:
        node_attrs = node_attributes(corpus, input_roots, normalize, order=order)
    if node_attrs.empty:
        return _pd.DataFrame(columns=["Root", "Spread", "Concentration",
                                      "Late-in-Ayah", "Late-in-Surah",
                                      "Mushaf Position", "Abundance"])
    total_ayahs = len(corpus.df)
    rows = []
    sprd_vals = node_attrs["Spread (Entropy)"].tolist()
    sprd_max = max(sprd_vals) if sprd_vals else 1
    burst_vals = node_attrs["Burstiness (Fano)"].tolist()
    burst_max = max(burst_vals) if burst_vals else 1
    log_max = math.log10(max(node_attrs["Total"].max(), 1) + 1) or 1
    pos_label = "Mushaf Position" if order == "mushaf" else "Rev-Order Position"
    for _, r in node_attrs.iterrows():
        rows.append({
            "Root": r["Root"],
            "Spread": round(r["Spread (Entropy)"] / sprd_max, 3) if sprd_max else 0,
            "Concentration": round(
                1 - (r["Burstiness (Fano)"] / burst_max), 3) if burst_max else 0,
            "Late-in-Ayah": round(r["Mean Pos-in-Ayah"], 3),
            "Late-in-Surah": round(r["Mean Pos-in-Surah"], 3),
            pos_label: round(r["Gravitational Center"] / max(total_ayahs - 1, 1), 3),
            "Abundance": round(math.log10(r["Total"] + 1) / log_max, 3),
        })
    return _pd.DataFrame(rows)


def meccan_medinan_pair_matrix(corpus, input_roots, normalize):
    """Per-pair table: in shared ayahs of A and B, what fraction are Meccan?
    Only meaningful if corpus.has_rev_order."""
    import pandas as _pd
    cols = ["Root A", "Root B", "Joint Ayahs",
            "Meccan Joint", "Medinan Joint", "Meccan % of Joint"]
    if not corpus.has_rev_order:
        return _pd.DataFrame(columns=cols)
    rev_lookup = corpus.rev_order_of_surah
    ayahs_of = {r: set(search_root(corpus, r, normalize)) for r in input_roots}
    rows = []
    for i in range(len(input_roots)):
        for j in range(i + 1, len(input_roots)):
            a, b = input_roots[i], input_roots[j]
            joint = ayahs_of[a] & ayahs_of[b]
            if not joint:
                continue
            meccan = medinan = 0
            for idx in joint:
                s = int(corpus.df.iloc[idx][COL_SURAH])
                ro = rev_lookup.get(s, 999)
                if ro <= MECCAN_CUTOFF: meccan += 1
                else: medinan += 1
            tot = meccan + medinan
            rows.append({
                "Root A": a, "Root B": b,
                "Joint Ayahs": tot,
                "Meccan Joint": meccan,
                "Medinan Joint": medinan,
                "Meccan % of Joint": round(100 * meccan / max(tot, 1), 1),
            })
    if not rows:
        return _pd.DataFrame(columns=cols)
    return _pd.DataFrame(rows).sort_values("Joint Ayahs", ascending=False)


def overlap_matrix_surah(corpus, input_roots, normalize=True):
    """Pairwise SURAH-level overlap: number of surahs in which BOTH appear."""
    import pandas as _pd
    K = (normalize_letters if normalize else (lambda t: t))
    surahs_with = {r: set() for r in input_roots}
    df = corpus.df
    for _, row in df.iterrows():
        sn = int(row[COL_SURAH])
        for tok in str(row[COL_ROOTS]).split():
            t = K(tok)
            for r in input_roots:
                if t == K(r):
                    surahs_with[r].add(sn)
    n = len(input_roots)
    m = _pd.DataFrame(0, index=input_roots, columns=input_roots, dtype=int)
    for i in range(n):
        for j in range(n):
            if i == j:
                m.iloc[i, j] = len(surahs_with[input_roots[i]])
            else:
                m.iloc[i, j] = len(surahs_with[input_roots[i]] &
                                   surahs_with[input_roots[j]])
    return m


# ===========================================================================
# GRAPH-NATIVE ANALYTICS for the Network tab — phase networks, directed
# lead-lag graph, per-root neighborhoods, network statistics
# ===========================================================================

def network_stats(g):
    """Comprehensive network-level statistics. Returns a dict suitable for
    a metric scoreboard at the top of the Network page."""
    if g is None or g.number_of_nodes() == 0:
        return {"nodes": 0, "edges": 0, "density": 0.0, "mean_degree": 0.0,
                "diameter": 0, "mean_shortest_path": 0.0,
                "modularity": 0.0, "assortativity": 0.0,
                "k_core_max": 0, "n_articulation_points": 0,
                "n_bridges": 0, "giant_component_pct": 0.0}
    n = g.number_of_nodes()
    e = g.number_of_edges()
    density = nx.density(g)
    mean_deg = (2 * e / n) if n else 0
    # Largest connected component
    if nx.is_connected(g):
        gc = g
        diameter = nx.diameter(g)
        mean_path = nx.average_shortest_path_length(g)
    else:
        components = list(nx.connected_components(g))
        gc_nodes = max(components, key=len)
        gc = g.subgraph(gc_nodes)
        try:
            diameter = nx.diameter(gc) if gc.number_of_nodes() > 1 else 0
            mean_path = (nx.average_shortest_path_length(gc)
                         if gc.number_of_nodes() > 1 else 0.0)
        except Exception:
            diameter = 0; mean_path = 0.0
    giant_pct = 100 * gc.number_of_nodes() / n
    # Modularity of the greedy community partition
    try:
        from networkx.algorithms.community import (greedy_modularity_communities,
                                                   modularity)
        comms = list(greedy_modularity_communities(g, weight="weight"))
        mod = modularity(g, comms, weight="weight") if comms else 0.0
    except Exception:
        mod = 0.0
    # Degree assortativity
    try:
        import warnings as _w, numpy as _np
        with _w.catch_warnings():
            _w.simplefilter("ignore")
            _old_err = _np.seterr(invalid="ignore", divide="ignore")
            try:
                assort = nx.degree_assortativity_coefficient(g)
                if assort is None or (isinstance(assort, float) and (_np.isnan(assort) or _np.isinf(assort))):
                    assort = 0.0
            finally:
                _np.seterr(**_old_err)
    except Exception:
        assort = 0.0
    # k-core
    try:
        cores = nx.core_number(g)
        k_core_max = max(cores.values()) if cores else 0
    except Exception:
        k_core_max = 0
    # Articulation points + bridges
    try: aps = list(nx.articulation_points(g))
    except Exception: aps = []
    try: bridges = list(nx.bridges(g))
    except Exception: bridges = []
    return {
        "nodes": n, "edges": e, "density": round(density, 4),
        "mean_degree": round(mean_deg, 2),
        "diameter": diameter,
        "mean_shortest_path": round(mean_path, 2),
        "modularity": round(mod, 3),
        "assortativity": round(assort, 3) if assort is not None else 0.0,
        "k_core_max": k_core_max,
        "n_articulation_points": len(aps),
        "n_bridges": len(bridges),
        "giant_component_pct": round(giant_pct, 1),
        "articulation_points": aps,
        "bridge_edges": bridges,
    }


def phase_networks(corpus, input_roots, normalize, top_partners=15, min_weight=1):
    """Build two co-occurrence networks separately from Meccan and Medinan
    ayahs. Returns (G_meccan, G_medinan). Only meaningful when the corpus
    carries revelation order."""
    if not corpus.has_rev_order:
        return None, None
    K = (normalize_letters if normalize else (lambda t: t))
    rev_lookup = corpus.rev_order_of_surah

    def _build_for(indices):
        partners = Counter()
        input_set = set(input_roots)
        for i in indices:
            roots_here = {K(t) for t in corpus.root_tokens[i]} - input_set
            for k in roots_here:
                partners[k] += 1
        keep = {p for p, _ in partners.most_common(top_partners)}
        node_set = set(input_roots) | keep
        edge_w = Counter()
        for i in indices:
            here = list({K(t) for t in corpus.root_tokens[i]} & node_set)
            for a, b in combinations(sorted(here), 2):
                edge_w[(a, b)] += 1
        g_ = nx.Graph()
        for nd in node_set:
            g_.add_node(nd, is_input=(nd in input_set))
        for (a, b), w in edge_w.items():
            if w >= min_weight:
                g_.add_edge(a, b, weight=w)
        return g_

    # Indices of ayahs that contain ANY input root, partitioned by phase
    match_ayahs = set()
    for q in input_roots:
        match_ayahs.update(search_root(corpus, q, normalize))
    meccan_idx, medinan_idx = [], []
    for i in match_ayahs:
        s = int(corpus.df.iloc[i][COL_SURAH])
        ro = rev_lookup.get(s, 999)
        if ro <= MECCAN_CUTOFF:
            meccan_idx.append(i)
        else:
            medinan_idx.append(i)
    return _build_for(meccan_idx), _build_for(medinan_idx)


def graph_diff(g1, g2):
    """Compute edge-level diff between two graphs: edges only in g1, only in
    g2, and in both. Each list is [(u, v, w1, w2), ...]."""
    if g1 is None: g1 = nx.Graph()
    if g2 is None: g2 = nx.Graph()
    edges1 = {tuple(sorted((u, v))): d.get("weight", 1) for u, v, d in g1.edges(data=True)}
    edges2 = {tuple(sorted((u, v))): d.get("weight", 1) for u, v, d in g2.edges(data=True)}
    both = sorted(set(edges1) & set(edges2))
    only1 = sorted(set(edges1) - set(edges2))
    only2 = sorted(set(edges2) - set(edges1))
    in_both = [(u, v, edges1[(u, v)], edges2[(u, v)]) for (u, v) in both]
    only_meccan = [(u, v, edges1[(u, v)], 0) for (u, v) in only1]
    only_medinan = [(u, v, 0, edges2[(u, v)]) for (u, v) in only2]
    return only_meccan, only_medinan, in_both


def directed_lead_lag_graph(corpus, input_roots, normalize, window=2,
                            min_strength=0.05):
    """Build a directed graph where edge A -> B exists if P(B within ±window
    ayahs of A | A) > min_strength AND this conditional is greater in the
    A->B direction than in the reverse. Weight = the conditional prob."""
    ll = lead_lag_matrix(corpus, input_roots, normalize, window=window)
    dg = nx.DiGraph()
    for r in input_roots:
        dg.add_node(r)
    for a in input_roots:
        for b in input_roots:
            if a == b: continue
            p_ab = float(ll.loc[a, b])
            p_ba = float(ll.loc[b, a])
            if p_ab > min_strength and p_ab > p_ba:
                dg.add_edge(a, b, weight=p_ab, asymmetry=round(p_ab - p_ba, 3))
    return dg


def per_root_neighborhood(g, root, depth=1, max_neighbors=8):
    """Return the ego-network around `root`: the root, its top
    max_neighbors strongest partners, and the edges among them."""
    if g is None or root not in g.nodes():
        return nx.Graph()
    # Pick top max_neighbors neighbors by edge weight
    nbrs = sorted(g.neighbors(root),
                  key=lambda n: g[root][n].get("weight", 1),
                  reverse=True)[:max_neighbors]
    nodes = {root} | set(nbrs)
    sub = g.subgraph(nodes).copy()
    return sub


def phase_filtered_node_attrs(corpus, input_roots, normalize, phase="Meccan"):
    """Recompute node_attributes restricted to ayahs of the given phase.
    Used to compare a concept's structural profile between Meccan and
    Medinan revelation."""
    import pandas as _pd
    if not corpus.has_rev_order:
        return _pd.DataFrame()
    rev_lookup = corpus.rev_order_of_surah
    K = (normalize_letters if normalize else (lambda t: t))
    rows = []
    for r in input_roots:
        rk = K(r)
        counts_by_surah = Counter()
        total = 0
        for i in search_root(corpus, r, normalize):
            s = int(corpus.df.iloc[i][COL_SURAH])
            ro = rev_lookup.get(s, 999)
            is_meccan = ro <= MECCAN_CUTOFF
            if (phase == "Meccan" and is_meccan) or (phase == "Medinan" and not is_meccan):
                counts_by_surah[s] += 1
                total += 1
        spread = _entropy(counts_by_surah.values())
        rows.append({
            "Root": r,
            f"{phase} Total": total,
            f"{phase} Surahs": len(counts_by_surah),
            f"{phase} Spread (Entropy)": round(spread, 3),
        })
    return _pd.DataFrame(rows)


def build_phase_subgraph(corpus, input_roots, normalize, rev_lo, rev_hi,
                         top_partners=15, min_weight=1):
    """Build a co-occurrence network using only ayahs whose surah has
    revelation-order in [rev_lo, rev_hi]. Used by the 4-stage evolution view."""
    if not corpus.has_rev_order:
        return None
    K = (normalize_letters if normalize else (lambda t: t))
    rev_lookup = corpus.rev_order_of_surah
    input_set = set(input_roots)
    # Indices in the phase window that contain any input root
    match_ayahs = set()
    for q in input_roots:
        for i in search_root(corpus, q, normalize):
            s = int(corpus.df.iloc[i][COL_SURAH])
            ro = rev_lookup.get(s, 999)
            if rev_lo <= ro <= rev_hi:
                match_ayahs.add(i)
    if not match_ayahs:
        return nx.Graph()
    partners = Counter()
    for i in match_ayahs:
        for k in {K(t) for t in corpus.root_tokens[i]} - input_set:
            partners[k] += 1
    keep = {p for p, _ in partners.most_common(top_partners)}
    node_set = input_set | keep
    edge_w = Counter()
    for i in match_ayahs:
        here = list({K(t) for t in corpus.root_tokens[i]} & node_set)
        for a, b in combinations(sorted(here), 2):
            edge_w[(a, b)] += 1
    g_ = nx.Graph()
    for nd in node_set:
        g_.add_node(nd, is_input=(nd in input_set))
    for (a, b), w in edge_w.items():
        if w >= min_weight:
            g_.add_edge(a, b, weight=w)
    return g_


# ---------------------------------------------------------------------------
# Per-Root Profile enhancements (v1.3): concentration stats + surface collocates
# ---------------------------------------------------------------------------
def root_concentration(corpus, q, normalize):
    """Gini, top-3 surah share, and breadth for a single root (frequency-side)."""
    import numpy as _np
    ayahs = search_root(corpus, q, normalize)
    by = Counter(int(corpus.df.iloc[i][COL_SURAH]) for i in ayahs)
    all_surahs = list(corpus.df[COL_SURAH].unique())
    counts = _np.array([by.get(int(s), 0) for s in all_surahs], dtype=float)
    tot = float(counts.sum())
    v = _np.sort(counts); n = len(v); gini = 0.0
    if n and v.sum() > 0:
        cum = _np.cumsum(v); gini = float((n + 1 - 2 * _np.sum(cum) / cum[-1]) / n)
    top3 = sum(c for _, c in by.most_common(3))
    return {"gini": round(gini, 3),
            "top3_share": round(100 * top3 / tot, 1) if tot else 0.0,
            "n_surahs": len(by)}


def surface_partner_lift(corpus, input_roots, normalize, top=15, min_co=4):
    """Top co-occurring SURFACE FORMS for each input root, ranked by LIFT.

    Lift = P(form | root-ayah) / P(form | any ayah). The root's OWN surface forms
    are excluded, and a support floor (>= min_co shared ayahs) suppresses unstable
    lift on rare forms. High lift = a word-form that travels with the root far more
    than its overall frequency predicts (a genuine collocate, not a ubiquitous word).
    """
    K = (normalize_letters if normalize else (lambda t: t))
    N = corpus.n_ayahs
    global_df = Counter()
    for stoks in corpus.surface_tokens:
        for s in set(stoks):
            global_df[s] += 1
    rows = []
    for q in input_roots:
        ayahs = search_root(corpus, q, normalize)
        n_in = len(ayahs)
        own = set()
        for i in ayahs:
            r_tokens = corpus.root_tokens[i]; s_tokens = corpus.surface_tokens[i]
            for j, t in enumerate(r_tokens):
                if K(t) == q and j < len(s_tokens):
                    own.add(s_tokens[j])
        co = Counter()
        for i in ayahs:
            for s in set(corpus.surface_tokens[i]):
                if s and s not in own:
                    co[s] += 1
        cand = []
        for s, nco in co.items():
            if nco < min_co:
                continue
            g = global_df.get(s, 0)
            if g <= 0 or n_in <= 0:
                continue
            lift = (nco / n_in) / (g / N)
            cand.append((s, nco, g, round(lift, 2), round(nco / n_in, 4)))
        cand.sort(key=lambda x: -x[3])
        for s, nco, g, lift, aff in cand[:top]:
            rows.append({"Input Root": q, "Partner Surface": s, "Ayahs Together": nco,
                         "Global Ayahs": g, "Lift": lift, "Affinity": aff})
    if not rows:
        return pd.DataFrame(columns=["Input Root", "Partner Surface", "Ayahs Together",
                                     "Global Ayahs", "Lift", "Affinity"])
    return pd.DataFrame(rows)
