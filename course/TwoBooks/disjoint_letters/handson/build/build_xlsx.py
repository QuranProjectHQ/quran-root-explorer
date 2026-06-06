import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
WK=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
B=json.load(open(os.path.join(WK,"dl_data_bank.json"),encoding="utf-8"))
wb=Workbook(); ws=wb.active; ws.title="Letter enrichment"
hdr=["#","letter","# bearer sūras","bearer-mean density","enrichment p","sample sūra","letter count","total letters","sample density %","verdict"]
ws.append(hdr)
for c0 in ws[1]: c0.font=Font(bold=True,color="FFFFFF"); c0.fill=PatternFill("solid",fgColor="305496")
for i,x in enumerate(B["letters"],1):
    v="enriched (artefact)" if x["p"]<0.05 else ("borderline" if x["p"]<0.12 else "no code")
    ws.append([i,x["letter"],x["n_bearers"],x["bearer_mean"],x["p"],x["sample_name"],x["sample_count"],x["sample_total"],x["sample_density_pct"],v])
for col,w in zip("ABCDEFGHIJ",[4,8,14,20,13,16,12,13,15,20]): ws.column_dimensions[col].width=w
ws2=wb.create_sheet("Validated finding")
for r in [["metric","value"],["ق density rank (Sūra 50)",B["qaf"]["sura50_rank"]],["ق enrichment p",B["qaf"]["p"]],
          ["contiguity p (muṣḥaf)",B["contiguity_p_mushaf"]],["median verses · muqaṭṭaʿāt",B["median_muq"]],
          ["median verses · others",B["median_other"]],["letters enriched / 14",B["n_sig"]]]:
    ws2.append(r)
ws2["A1"].font=ws2["B1"].font=Font(bold=True)
wb.save(os.path.join(WK,"DisjointLetters_Data_Bank.xlsx")); print("xlsx built")
