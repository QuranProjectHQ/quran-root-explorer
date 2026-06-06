# -*- coding: utf-8 -*-
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
HEAD=Font(name='Arial',bold=True,color='FFFFFF',size=11); BODY=Font(name='Arial',size=11)
FILL=PatternFill('solid',start_color='1E2761'); CEN=Alignment(horizontal='center',vertical='center')
THIN=Side(style='thin',color='D0D0D0'); BORD=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
def sheet(wb,title,headers,rows,widths=None,first=False):
    ws=wb.active if first else wb.create_sheet(); ws.title=title
    ws.append(headers)
    for c in ws[1]:
        c.font=HEAD; c.fill=FILL; c.alignment=CEN; c.border=BORD
    for r in rows:
        ws.append(r)
        for c in ws[ws.max_row]:
            c.font=BODY; c.border=BORD; c.alignment=Alignment(vertical='center')
    ws.freeze_panes='A2'
    for i,w in enumerate(widths or [16]*len(headers),1):
        ws.column_dimensions[ws.cell(row=1,column=i).column_letter].width=w
    return ws

# ===================== WEEK 2 =====================
k=json.load(open('week02/wk2_keys.json')); mk=json.load(open('week02/wk2_member_keys.json')); ec=json.load(open('week02/wk2_exercise_calc.json'))
wb=Workbook()
# Profiles
rows=[]
for rt,p in k['profiles'].items():
    hr=p['home_raw']; hs=p['home_size_true']
    rows.append([rt,p['freq_ayahs'],p['breadth_surahs'],f"{p['top3_share']}%",p['gini'],
                 f"{hr['name']} ({hr['ayah_hits']})",
                 (f"{hs['name']} ({hs['per_1k_roots']}/1k rt)" if isinstance(hs,dict) else hs)])
sheet(wb,'Profiles',['root','freq (ayahs)','breadth /114','top-3 share','Gini','raw busiest surah','size-true home (per 1k root-tokens)'],rows,[10,12,12,11,8,22,34],first=True)
# Member assignments
rows=[]
for rt,m in mk.items():
    hs=m['home_size_true']
    rows.append([m['member'],rt,m['freq_ayahs'],m['breadth'],f"{m['top3']}%",m['gini'],
                 f"{m['home_raw'][0]} ({m['home_raw'][1]})",
                 (f"{hs[0]} ({hs[1]}/1k rt)" if isinstance(hs,list) else hs)])
rows.sort(key=lambda r:r[0])
sheet(wb,'Member assignments',['#','root','freq','breadth','top-3','Gini','raw busiest','size-true home'],rows,[5,10,8,10,9,8,20,26])
# Exercise Part-1 calc
rows=[]
for rt,e in ec.items():
    r=e['raw']; h=e['home']
    rows.append([rt,f"{r[0]}",f"{r[1]}/{r[2]}",f"{r[3]}",f"{h[0]}",f"{h[1]}/{h[2]}",f"{h[3]}"])
sheet(wb,'Exercise Part-1 calc',['root','raw surah','tok/surah-tok','per-1k rt','size-true home','tok/surah-tok','per-1k rt'],rows,[10,16,16,11,16,16,11])
# Meta
m=k['meta']; h=m['headline']
metarows=[['n_ayahs',m['n_ayahs']],['n_surahs',m['n_surahs']],['n_root_tokens',m['n_root_tokens']],
['largest surah',f"{m['largest_surah']['name']} ({m['largest_surah']['ayahs']} ayahs, {m['largest_surah']['root_tokens']} tokens)"],
['size-true unit',m['size_true_unit']],['floor',f"count ≥ {m['floor']['min_count']} AND surah ≥ {m['floor']['min_surah_root_tokens']} root-tokens"],
['HEADLINE — raw busiest = al-Baqara',f"{h['raw_home_albaqara']} of top {h['top_n']}"],
['HEADLINE — per-ayah home = al-Baqara',f"{h['per_ayah_home_albaqara']} of {h['top_n']}"],
['HEADLINE — per-root-tokens home = al-Baqara',f"{h['per_roots_home_albaqara']} of {h['top_n']}"]]
sheet(wb,'Meta & headline',['key','value'],metarows,[42,40])
wb.save('week02/Week2_Data_Bank.xlsx'); print('Week2_Data_Bank.xlsx saved')

# ===================== WEEK 3 =====================
k=json.load(open('week03/wk3_keys.json'))
wb=Workbook()
rows=[[rt,k['gloss'].get(rt,''),k['freq'][rt],k['nforms'][rt]] for rt in k['freq']]
rows.sort(key=lambda r:-r[3])
sheet(wb,'Roots — freq & richness',['root','gloss','freq (ayahs)','# distinct forms'],rows,[10,16,12,16],first=True)
# Forms of worked root amn
rows=[[f,c,f"{c/sum(x[1] for x in k['forms']['ءمن'])*100:.1f}%"] for f,c in k['forms']['ءمن']]
sheet(wb,'Forms — amن (worked)',['surface form','count','share'],rows,[14,10,10])
# Pattern families
rows=[[fam,v[0],f"{v[1]}%"] for fam,v in k['amn_families'].items()]
sheet(wb,'Pattern families (amن)',['pattern family','tokens','share'],rows,[24,10,10])
# Divine names
attr={'رحيم':'the All-Merciful','حكيم':'the All-Wise','غفور':'the Oft-Forgiving','رحمن':'the Most Gracious','سميع':'the All-Hearing','بصير':'the All-Seeing'}
rows=[[n,r,c,attr.get(n,'')] for n,c,r in sorted(k['divine'],key=lambda x:-x[1])]
sheet(wb,'Divine Names (intensives)',['name','root','occurrences','attribute'],rows,[12,10,12,20])
# Partners
rows=[[b,j,l,z] for b,j,l,z in k['partners_amn']]
sheet(wb,'Partners (amن)',['partner root','joint ayahs','lift (×)','significance z'],rows,[14,12,10,14])
# Antonyms
rows=[[f"{a} ↔ {b}",j] for a,b,j in k['antonyms']]
sheet(wb,'Antonym partners',['pair','shared ayahs'],rows,[18,14])
wb.save('week03/Week3_Data_Bank.xlsx'); print('Week3_Data_Bank.xlsx saved')
